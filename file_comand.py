import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, colors
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import IconSetRule

st.subheader("📋 Elaborare file excel", divider='blue')

data_clienti = {
    'ID_cliente': [1, 2, 3, 4, 5, 6, 7],
    'nome_cliente': ['Mario Rossi', 'Laura Bianchi', 'Giuseppe Verdi', 'Anna Neri', 'Paolo Forti', 'Silvia Gallo', 'Marco Bini'],
    'citta': ['Roma', 'Milano', 'Napoli', 'Torino', 'Bologna', 'Milano', 'Roma']
}
clienti = pd.DataFrame(data_clienti)

data_ordini = {
    'ID_ordine': [101, 102, 103, 104, 105, 106, 107, 108],
    'ID_cliente': [1, 3, 2, 1, 5, 6, 10, 4],
    'data_ordine': pd.to_datetime(['2024-01-15', '2024-01-17', '2024-02-01', '2024-02-05', '2024-02-10', '2024-03-01', '2024-03-05', '2024-03-10']),
    'Prodotto': ['Telecaster', 'Stratocaster', 'Les Paul', 'SG', 'Jazzmaster','Es-335','Gretsch White Falcon','Flying V'],
    'importo': [150.50, 200.00, 75.25, 300.00, 120.00, 80.00, 50.00, 250.00]
}
ordini = pd.DataFrame(data_ordini)

with st.expander("**:inbox_tray: Import**"):
	code = '''
	from io import BytesIO
	from openpyxl import load_workbook, Workbook
	import os
	'''
	st.code(code, language="python")
#************************************************************************************************************************************************************************************
with st.expander("**:open_file_folder: Leggere un file .csv e .xlsx**"):
    st.write("#### 1. Lettura file csv")
    code = '''
    file_csv = pd.read_csv(percorso_del_file_csv, sep = ';')
    dataframe = pd.DataFrame(file_csv)'''
    st.code(code, language="python")
    st.write("#### 2. Lettura file excel")
    code = '''
    file_excel = pd.read_excel(percorso_del_file_excel, sheet_name= 'FoglioXXX', usecols=['Pippo', 'Pluto', 'Team', 'Ingresso']) 
    dataframe = pd.DataFrame(file_excel)'''
    st.code(code, language="python")
    st.info("usecols è opzionale e puoi decidere quali colonne importare, se non viene inserito verrà importato tutto il file")

#************************************************************************************************************************************************************************************
with st.expander("**:arrows_counterclockwise: Gestire Colonne Dinamiche con Mappatura**"):
    st.write("Questa sezione mostra come creare un'applicazione flessibile per analizzare file in cui l'ordine o il nome delle colonne può cambiare.")
    
    st.subheader("1. Carica il tuo file di report")
    code = '''uploaded_file = st.file_uploader(
    "Scegli un file Excel o CSV",
    type=['csv', 'xlsx'],
    key="file_uploader_dinamico"
)'''
    st.code(code, language="python")
    uploaded_file = st.file_uploader(
        "Scegli un file Excel o CSV",
        type=['csv', 'xlsx'],
        key="file_uploader_dinamico"
    )

    df_originale = None
    if uploaded_file:
        try:
            if uploaded_file.name.endswith('.csv'):
                first_line = uploaded_file.readline().decode('utf-8')
                dialect = csv.Sniffer().sniff(first_line)
                separatore_rilevato = dialect.delimiter
                st.info(f"Separatore del CSV rilevato automaticamente: **'{separatore_rilevato}'**")
                uploaded_file.seek(0)
                df_originale = pd.read_csv(uploaded_file, sep=separatore_rilevato)
            else:
                df_originale = pd.read_excel(uploaded_file)
        except Exception as e:
            st.error(f"Si è verificato un errore durante la lettura del file: {e}")
            df_originale = None # Assicura che non si proceda se c'è errore

    if df_originale is not None:
        st.success("File caricato con successo!")
        st.write("🔍 **Anteprima dei dati caricati**:")
        st.dataframe(df_originale.head())

        st.subheader("2. Mappa le colonne del tuo file 🗺️")
        
        colonne_richieste = {
            'ID Ordine': 'L\'identificativo univoco dell\'ordine.',
            'Data Ordine': 'La data in cui è stato effettuato l\'ordine.',
            'Prodotto': 'Il nome del prodotto venduto.',
            'Importo': 'L\'importo numerico dell\'ordine.'
        }
        mappatura_colonne = {}
        opzioni_colonne = df_originale.columns.tolist()
        for nome_logico, descrizione in colonne_richieste.items():
            mappatura_colonne[nome_logico] = st.selectbox(
                f"Quale colonna rappresenta: **{nome_logico}**?",
                options=opzioni_colonne,
                help=descrizione,
                key=f"map_{nome_logico}"
            )

        code_map = '''colonne_richieste = {
    'ID Ordine': "L\'identificativo univoco dell\'ordine.",
    'Data Ordine': "La data in cui è stato effettuato l\'ordine.",
    'Prodotto': "Il nome del prodotto venduto.",
    'Importo': "L\'importo numerico dell\'ordine."
}
mappatura_colonne = {}
opzioni_colonne = df_originale.columns.tolist()
for nome_logico, descrizione in colonne_richieste.items():
    mappatura_colonne[nome_logico] = st.selectbox(
        f"Quale colonna rappresenta: **{nome_logico}**?",
        options=opzioni_colonne,
        help=descrizione,
        key=f"map_{nome_logico}"
    )'''
        st.code(code_map, language="python")

        st.subheader("3. Creazione del Report Standardizzato ✅")

        code_standard = '''dict_rinomina = {valore_scelto: nome_logico for nome_logico, valore_scelto in mappatura_colonne.items()}
df_standard = df_originale.rename(columns=dict_rinomina)
df_standard = df_standard[list(colonne_richieste.keys())]
st.dataframe(df_standard)'''
        st.code(code_standard, language="python")

        if st.button("Crea Report Standardizzato"):
            dict_rinomina = {valore_scelto: nome_logico for nome_logico, valore_scelto in mappatura_colonne.items()}
            df_standard = df_originale.rename(columns=dict_rinomina)
            df_standard = df_standard[list(colonne_richieste.keys())]
            st.write("**Ecco il tuo DataFrame pulito e standardizzato, pronto per l'analisi!**")
            st.dataframe(df_standard)


#************************************************************************************************************************************************************************************    
with st.expander("**:outbox_tray: Esportare un DataFrame in Excel**"):

    tab1, tab2 = st.tabs(["**Standar**", "**Con formattazione**"])
    
    with tab1:

        st.write(
            """
            Una volta che i dati sono stati elaborati, è spesso necessario esportarli in un formato
            come Excel. Streamlit rende questo processo semplice con il `st.download_button`.
            
            Il trucco consiste nel creare il file Excel "in memoria" (usando `BytesIO`) invece che 
            su un file fisico sul disco, per poi passarlo al pulsante di download.
            """
        )

        # Usiamo il DataFrame 'clienti' e 'ordini' come esempio
        st.write("Esempio di esportazione dei DataFrame 'clienti' e 'ordini' in un unico file Excel con due fogli di lavoro e l'aggiunta di layout personalizzati")

        code = '''
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = BytesIO()

    # Usiamo pd.ExcelWriter per scrivere più fogli nello stesso file
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        clienti.to_excel(writer, index=False, sheet_name='Clienti')
        ordini.to_excel(writer, index=False, sheet_name='Ordini')

        # --- Accesso al workbook e ai fogli di lavoro di openpyxl ---
        workbook = writer.book
        worksheet_clienti = writer.sheets['Clienti']
        worksheet_ordini = writer.sheets['Ordini']

    # Recuperiamo i dati binari del file Excel creato in memoria
    file_data = output.getvalue()

    st.download_button(
        label="📥 Scarica File Excel Standard",
        data=file_data,
        file_name="Export_Formattato.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
        '''
        st.code(code, language='python')
        
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            clienti.to_excel(writer, index=False, sheet_name='Clienti')
            ordini.to_excel(writer, index=False, sheet_name='Ordini')

            workbook = writer.book
            worksheet_clienti = writer.sheets['Clienti']
            worksheet_ordini = writer.sheets['Ordini']

        # Recuperiamo i dati binari del file Excel creato in memoria
        file_data = output.getvalue()

        st.download_button(
            label="📥 Scarica File Excel Standard",
            data=file_data,
            file_name="Export_Formattato.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with tab2:
        st.markdown("""
        Questa sezione mostra come esportare uno o più DataFrame in un singolo file Excel, applicando stili personalizzati alle celle.
        Il processo chiave è:
        1.  Creare un file Excel in memoria con `BytesIO`.
        2.  Usare `pandas.ExcelWriter` con il motore `openpyxl` per scrivere i dati.
        3.  Accedere al `workbook` e ai `worksheet` di openpyxl attraverso l'oggetto `writer`.
        4.  Definire e applicare gli stili (font, colori, bordi, allineamento).
        5.  Aggiungere formattazione condizionale (es. icone).
        6.  Passare i dati binari del file a `st.download_button`.
        """)

        # --- Codice da visualizzare ---
        code_export = '''
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, colors
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import IconSetRule

    # 1. Creazione del file in memoria
    output = BytesIO()

    # 2. Scrittura con pd.ExcelWriter
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        clienti.to_excel(writer, index=False, sheet_name='Clienti')
        ordini.to_excel(writer, index=False, sheet_name='Ordini')

        # 3. Accesso agli oggetti openpyxl
        workbook = writer.book
        ws_clienti = writer.sheets['Clienti']
        ws_ordini = writer.sheets['Ordini']

        # 4. Definizione degli stili
        header_font = Font(bold=True, color="FFFFFF", name="Calibri")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Funzione helper per applicare stile
        def format_sheet(worksheet):
            # Stile intestazione
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Adatta larghezza colonne e applica bordi
            for col in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    cell.border = thin_border
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # 5. Applicazione degli stili
        format_sheet(ws_clienti)
        format_sheet(ws_ordini)

        # 6. Aggiunta Formattazione Condizionale (IconSet)
        # Mostra icone a semaforo per la colonna 'importo' nel foglio 'Ordini'
        icon_set_rule = IconSetRule(
            '3TrafficLights1', # Stile delle icone
            'num', # Tipo di dato
            [1000, 3000, None], # Valori soglia: verde > 3000, giallo > 1000, rosso <= 1000
            showValue=None, 
            percent=None
        )
        # Applica la regola all'intervallo di celle corretto (es. E2:E9)
        ws_ordini.conditional_formatting.add(f'E2:E{ws_ordini.max_row}', icon_set_rule)


    # 7. Ottenimento dati binari per il download
    file_data = output.getvalue()

    st.download_button(
        label="📥 Scarica File Excel Formattato",
        data=file_data,
        file_name="Report_Avanzato.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
        '''
        st.code(code_export, language='python')

        # --- Esecuzione del codice per il download ---
        output_buffer = BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            clienti.to_excel(writer, index=False, sheet_name='Clienti')
            ordini.to_excel(writer, index=False, sheet_name='Ordini')

            ws_clienti = writer.sheets['Clienti']
            ws_ordini = writer.sheets['Ordini']

            header_font = Font(bold=True, color="FFFFFF", name="Calibri")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            def _format_sheet(worksheet):
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                for col in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(col[0].column)
                    for cell in col:
                        cell.border = thin_border
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            _format_sheet(ws_clienti)
            _format_sheet(ws_ordini)

            icon_set_rule = IconSetRule('3TrafficLights1', 'num', [1000, 3000, None])
            ws_ordini.conditional_formatting.add(f'E2:E{ws_ordini.max_row}', icon_set_rule)

        excel_data = output_buffer.getvalue()
        st.download_button(
            label="📥 Scarica File Excel Formattato",
            data=excel_data,
            file_name="Report_Avanzato.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

